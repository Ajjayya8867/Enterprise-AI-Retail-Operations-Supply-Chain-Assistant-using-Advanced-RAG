import os
import csv
import re
import zipfile
import xml.etree.ElementTree as ET
from pypdf import PdfReader

# Optional external imports with clean fallbacks
try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def clean_text(text):
    # Remove excessive whitespace
    text = re.sub(r'[ \t\r\f\v]+', ' ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()

def recursive_chunk_text(text, max_size=800, overlap=150):
    """
    Splits text recursively using structural boundaries:
    1. Paragraphs (\n\n)
    2. Newlines (\n)
    3. Sentences (. )
    4. Words ( )
    """
    text = clean_text(text)
    if not text:
        return []
        
    paragraphs = text.split("\n\n")
    chunks = []
    current_chunk = []
    current_length = 0
    
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
            
        para_words = para.split(" ")
        # If paragraph fits within max_size, add it
        if len(para_words) + current_length <= max_size:
            current_chunk.append(para)
            current_length += len(para_words)
        else:
            # Paragraph doesn't fit, commit current chunk if it has text
            if current_chunk:
                chunks.append("\n\n".join(current_chunk))
                # Create overlap: keep last N words
                last_para = current_chunk[-1].split(" ")
                overlap_words = last_para[-min(len(last_para), overlap):]
                current_chunk = [" ".join(overlap_words)]
                current_length = len(overlap_words)
            
            # If a single paragraph is too large, split it by sentences
            if len(para_words) > max_size:
                sentences = re.split(r'(?<=[.!?])\s+', para)
                for sentence in sentences:
                    sentence_words = sentence.split(" ")
                    if len(sentence_words) + current_length <= max_size:
                        current_chunk.append(sentence)
                        current_length += len(sentence_words)
                    else:
                        if current_chunk:
                            chunks.append("\n\n".join(current_chunk))
                            # overlap
                            last_sent = current_chunk[-1].split(" ")
                            overlap_words = last_sent[-min(len(last_sent), overlap):]
                            current_chunk = [" ".join(overlap_words)]
                            current_length = len(overlap_words)
                        
                        # If a single sentence is still too large, split it by words
                        if len(sentence_words) > max_size:
                            for i in range(0, len(sentence_words), max_size - overlap):
                                sub_sentence = " ".join(sentence_words[i:i+max_size])
                                chunks.append(sub_sentence)
                            current_chunk = []
                            current_length = 0
                        else:
                            current_chunk.append(sentence)
                            current_length += len(sentence_words)
            else:
                current_chunk.append(para)
                current_length += len(para_words)
                
    if current_chunk:
        chunks.append("\n\n".join(current_chunk))
        
    return chunks

def build_parent_child_chunks(text, source_name, doc_type, parent_size=800, child_size=150, overlap=30):
    """
    Implements Parent-Child chunking:
    - Splits text into large 'parent' sections.
    - Splits each parent section into smaller 'child' chunks.
    - The child chunk contains a pointer to the parent section text in its metadata.
    """
    text = clean_text(text)
    if not text:
        return []
        
    parents = recursive_chunk_text(text, max_size=parent_size, overlap=parent_size // 4)
    chunks = []
    
    for p_idx, parent_text in enumerate(parents):
        parent_words = parent_text.split(" ")
        step = child_size - overlap
        if step <= 0:
            step = child_size // 2
            
        for c_idx in range(0, len(parent_words), step):
            child_words = parent_words[c_idx:c_idx + child_size]
            child_text = " ".join(child_words)
            
            # Skip very short trailing child chunks
            if len(child_text) < 40 and len(parent_words) > child_size:
                continue
                
            chunks.append({
                "text": child_text,
                "metadata": {
                    "source": source_name,
                    "type": doc_type,
                    "parent_text": parent_text,
                    "parent_index": p_idx,
                    "child_index": c_idx // step,
                    "start_word": c_idx,
                    "end_word": c_idx + len(child_words)
                }
            })
    return chunks

def parse_txt(file_path):
    source_name = os.path.basename(file_path)
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            text = f.read()
        return build_parent_child_chunks(text, source_name, "txt")
    except Exception as e:
        print(f"Error parsing TXT {file_path}: {e}")
        return []

def parse_pdf(file_path):
    source_name = os.path.basename(file_path)
    try:
        reader = PdfReader(file_path)
        full_text = ""
        page_mappings = []
        current_word_count = 0
        
        for page_idx, page in enumerate(reader.pages):
            page_text = page.extract_text() or ""
            page_text_clean = clean_text(page_text)
            if page_text_clean:
                full_text += page_text_clean + "\n\n"
                words_in_page = len(page_text_clean.split(' '))
                page_mappings.append({
                    "page": page_idx + 1,
                    "start_word": current_word_count,
                    "end_word": current_word_count + words_in_page
                })
                current_word_count += words_in_page
                
        # Generate parent-child chunks
        raw_chunks = build_parent_child_chunks(full_text, source_name, "pdf")
        
        # Link child chunks back to page numbers based on start_word
        for chunk in raw_chunks:
            start = chunk["metadata"]["start_word"]
            end = chunk["metadata"]["end_word"]
            
            pages = []
            for pm in page_mappings:
                if (start >= pm["start_word"] and start <= pm["end_word"]) or \
                   (end >= pm["start_word"] and end <= pm["end_word"]) or \
                   (start <= pm["start_word"] and end >= pm["end_word"]):
                    pages.append(pm["page"])
            
            chunk["metadata"]["pages"] = pages if pages else [1]
            
        return raw_chunks
    except Exception as e:
        print(f"Error parsing PDF {file_path}: {e}")
        return []

def parse_docx_fallback(file_path):
    """Zero-dependency docx loader extracting text directly from document.xml zip file."""
    try:
        with zipfile.ZipFile(file_path) as z:
            xml_content = z.read('word/document.xml')
            root = ET.fromstring(xml_content)
            # Namespace mapping
            ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            paragraphs = []
            for p in root.findall('.//w:p', ns):
                texts = [t.text for t in p.findall('.//w:t', ns) if t.text]
                if texts:
                    paragraphs.append("".join(texts))
            return "\n\n".join(paragraphs)
    except Exception as e:
        print(f"Fallback docx parser failed: {e}")
        return ""

def parse_docx(file_path):
    source_name = os.path.basename(file_path)
    text = ""
    if HAS_DOCX:
        try:
            doc = docx.Document(file_path)
            text = "\n\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        except Exception as e:
            print(f"python-docx failed: {e}. Using XML fallback.")
            text = parse_docx_fallback(file_path)
    else:
        text = parse_docx_fallback(file_path)
        
    return build_parent_child_chunks(text, source_name, "docx")

def parse_xlsx(file_path):
    source_name = os.path.basename(file_path)
    rows_text = []
    
    if HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_vals = [str(val).strip() for val in row if val is not None]
                    if row_vals:
                        rows_text.append(f"In Sheet '{sheet_name}', row #{r_idx+1}: " + ", ".join(row_vals))
            text = "\n".join(rows_text)
            return build_parent_child_chunks(text, source_name, "xlsx")
        except Exception as e:
            print(f"openpyxl failed: {e}. Fallback to basic text parse.")
            
    # Basic text fallback for binary XLSX
    return parse_txt(file_path)

def parse_csv(file_path):
    source_name = os.path.basename(file_path)
    chunks = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            sample = f.read(2048)
            f.seek(0)
            delimiter = ';' if ';' in sample and sample.count(';') > sample.count(',') else ','
            
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames
            
            for row_idx, row in enumerate(reader):
                row_descriptions = []
                for col in headers:
                    val = row.get(col, '')
                    if val is not None and val.strip() != '':
                        row_descriptions.append(f"{col}: {val.strip()}")
                
                row_text = f"In document '{source_name}', record #{row_idx + 1} contains: " + ", ".join(row_descriptions)
                
                # For CSV rows, the child chunk is the row, and the parent is also the row or surrounding rows.
                chunks.append({
                    "text": row_text,
                    "metadata": {
                        "source": source_name,
                        "type": "csv",
                        "parent_text": row_text,
                        "row_index": row_idx + 1,
                        "data": dict(row)
                    }
                })
        return chunks
    except Exception as e:
        print(f"Error parsing CSV {file_path}: {e}")
        return []

def parse_file(file_path):
    _, ext = os.path.splitext(file_path.lower())
    if ext == '.txt':
        return parse_txt(file_path)
    elif ext == '.pdf':
        return parse_pdf(file_path)
    elif ext == '.csv':
        return parse_csv(file_path)
    elif ext in ['.docx', '.doc']:
        return parse_docx(file_path)
    elif ext in ['.xlsx', '.xls']:
        return parse_xlsx(file_path)
    else:
        # Fallback
        try:
            return parse_txt(file_path)
        except:
            return []
